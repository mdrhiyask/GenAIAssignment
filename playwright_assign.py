"""WhatsApp Web automation assignment. Requires: pip install playwright openpyxl; playwright install."""
from __future__ import annotations

import json
import random
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from playwright.sync_api import Locator, Page, TimeoutError as PlaywrightTimeoutError, sync_playwright

BASE_DIR = Path(__file__).resolve().parent
CONTACTS_FILE = BASE_DIR / "contacts.xlsx"
SESSION_DIR = BASE_DIR / "whatsapp_session"  # Reuses the WhatsApp login on later runs.
SCREENSHOTS_DIR = BASE_DIR / "screenshots"
DELAY_MS = (2_000, 5_000)


def pause(page: Page) -> None:
    page.wait_for_timeout(random.randint(*DELAY_MS))


def phone_digits(value: Any) -> str:
    return re.sub(r"\D", "", str(value or ""))


def visible(page: Page, selectors: list[str], timeout: int = 7_000) -> Locator:
    """Try alternative selectors because WhatsApp Web's UI can change."""
    last_error: Exception | None = None
    for selector in selectors:
        locator = page.locator(selector).first
        try:
            locator.wait_for(state="visible", timeout=timeout)
            return locator
        except PlaywrightTimeoutError as error:
            last_error = error
    raise PlaywrightTimeoutError(f"No selector became visible: {selectors}") from last_error


def read_contacts() -> list[dict[str, str]]:
    if not CONTACTS_FILE.exists():
        raise FileNotFoundError("Create contacts.xlsx next to this script with Name, Phone and Message columns.")
    book = load_workbook(CONTACTS_FILE, read_only=True, data_only=True)
    sheet = book.active
    header_row = next(sheet.iter_rows(values_only=True))
    headers = {str(value).strip().lower(): i for i, value in enumerate(header_row) if value is not None}
    missing = {"name", "phone"} - headers.keys()
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")
    contacts = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, phone = str(row[headers["name"]] or "").strip(), phone_digits(row[headers["phone"]])
        template = str(row[headers["message"]] or "Hello {name}!").strip() if "message" in headers else "Hello {name}!"
        if name and phone:
            contacts.append({"name": name, "phone": phone, "template": template})
    return contacts


def login(page: Page) -> None:
    page.goto("https://web.whatsapp.com/", wait_until="domcontentloaded")
    print("Scan the QR code on the first run, if WhatsApp asks for it.")

    try:
        # #side is the left chat panel, visible after WhatsApp Web has loaded.
        page.wait_for_selector("#side", state="visible", timeout=180_000)
        page.wait_for_timeout(2_000)
    except PlaywrightTimeoutError as error:
        page.screenshot(path=str(BASE_DIR / "whatsapp_login_error.png"))
        raise RuntimeError(
            "WhatsApp Web did not finish loading after QR scan. "
            "Check your internet connection and make sure WhatsApp Web is open."
        ) from error


def open_chat(page: Page, phone: str) -> None:
    page.goto(f"https://web.whatsapp.com/send?phone={phone}", wait_until="domcontentloaded")
    try:
        visible(page, ['footer div[contenteditable="true"]', 'div[contenteditable="true"][data-tab="10"]'], 20_000)
    except PlaywrightTimeoutError as error:
        raise RuntimeError("Contact was not found or the chat could not be opened.") from error


def send(page: Page, text: str) -> None:
    composer = visible(page, ['footer div[contenteditable="true"][data-tab="10"]', 'footer div[contenteditable="true"]', 'div[contenteditable="true"][aria-label="Type a message"]'])
    composer.click()
    composer.fill(text)
    pause(page)
    composer.press("Enter")
    page.locator("div.message-out").last.wait_for(state="visible", timeout=10_000)


def incoming_messages(page: Page) -> list[str]:
    bubbles = page.locator("div.message-in span.selectable-text")
    return [bubbles.nth(i).inner_text().strip() for i in range(max(0, bubbles.count() - 3), bubbles.count()) if bubbles.nth(i).inner_text().strip()]


def reports(results: list[dict[str, Any]]) -> tuple[Path, Path]:
    stamp = date.today().isoformat()
    json_file = BASE_DIR / f"whatsapp_report_{stamp}.json"
    excel_file = BASE_DIR / f"whatsapp_report_{stamp}.xlsx"
    json_file.write_text(json.dumps(results, indent=2, ensure_ascii=False), encoding="utf-8")
    book = Workbook()
    sheet = book.active
    sheet.title = "WhatsApp Report"
    sheet.append(["Name", "Phone", "Status", "Sent At", "Message", "Last 3 Incoming Messages", "Screenshot", "Error"])
    for item in results:
        sheet.append([item["name"], item["phone"], item["status"], item["sent_at"], item["message"], "\n".join(item["last_3_messages"]), item["screenshot"], item["error"]])
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 55)
    book.save(excel_file)
    return json_file, excel_file


def main() -> None:
    contacts = read_contacts()
    if not contacts:
        raise ValueError("contacts.xlsx has no usable contacts.")
    SCREENSHOTS_DIR.mkdir(exist_ok=True)
    results: list[dict[str, Any]] = []
    with sync_playwright() as playwright:
        context = playwright.chromium.launch_persistent_context(str(SESSION_DIR), headless=False, viewport={"width": 1280, "height": 900})
        page = context.pages[0] if context.pages else context.new_page()
        try:
            login(page)
            for contact in contacts:
                message = contact["template"].replace("{name}", contact["name"])
                item: dict[str, Any] = {"name": contact["name"], "phone": contact["phone"], "message": message, "status": "failed", "sent_at": "", "last_3_messages": [], "screenshot": "", "error": ""}
                try:
                    open_chat(page, contact["phone"])
                    pause(page)
                    send(page, message)
                    item["status"], item["sent_at"] = "sent", datetime.now().isoformat(timespec="seconds")
                    item["last_3_messages"] = incoming_messages(page)
                    image = SCREENSHOTS_DIR / f"{re.sub(r'[^A-Za-z0-9_-]+', '_', contact['name'])}_{contact['phone']}.png"
                    page.screenshot(path=str(image))
                    item["screenshot"] = str(image)
                    print(f"Sent to {contact['name']}.")
                except Exception as error:
                    item["error"] = str(error)
                    print(f"Skipped {contact['name']}: {error}")
                results.append(item)
                pause(page)
        finally:
            context.close()
    json_file, excel_file = reports(results)
    print(f"Reports saved: {json_file.name}, {excel_file.name}")


if __name__ == "__main__":
    main()
